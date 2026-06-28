#!/usr/bin/env python3

from pathlib import Path
import argparse

import pandas as pd
from openpyxl import load_workbook
from io import StringIO


COLUMN_MAP = {
    "last_price": ["Last price"],
    "bid": ["Bid"],
    "ask": ["Ask"],
    "quantity": ["Quantity", "Qty"],
    "market_value": ["Current value"],
    "symbol": ["Symbol"],
}


def build_lookup(columns) -> dict[str, str]:
    return {column.lower().strip(): column for column in columns}


def get_column(lookup: dict[str, str], logical_name: str) -> str | None:
    for candidate in COLUMN_MAP.get(logical_name, []):
        column = lookup.get(candidate.lower())
        if column:
            return column
    return None


def parse_args():
    parser = argparse.ArgumentParser(
        description="Calculate midpoint values for a Fidelity options portfolio."
    )

    parser.add_argument(
        "input_csv",
        type=Path,
        help="Input Fidelity CSV file",
    )

    parser.add_argument(
        "output_xlsx",
        nargs="?",
        type=Path,
        help="Output Excel workbook (default: input filename with .xlsx extension)",
    )

    return parser.parse_args()


def format_currency(ws, headers, column_name):
    if column_name and column_name in headers:
        col = headers[column_name]
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = "$#,##0.00"


def main() -> int:
    args = parse_args()

    input_path = args.input_csv

    if not input_path.exists():
        print(f"Error: '{input_path}' does not exist.")
        return 1

    output_path = (
        args.output_xlsx
        if args.output_xlsx
        else input_path.with_suffix(".xlsx")
    )

    if output_path.exists():
        answer = input(
            f"'{output_path}' already exists. Overwrite? [Y/n]: "
        ).strip().lower()

        if answer in ("n", "no"):
            print("Aborted.")
            return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)

    #
    # Read CSV
    #

    with open(input_path, encoding="utf-8-sig") as f:
        lines = f.readlines()

    # Fidelity's disclaimer marks the beginning of the footer.
    for i, line in enumerate(lines):
        if line.startswith('"The data and information'):
            lines = lines[:i]
            break

    # Fix Fidelity's missing trailing comma in the header.
    header = lines[0].rstrip("\n")
    if header.count(",") < lines[1].count(","):
        lines[0] = header + ",\n"

    df = pd.read_csv(StringIO("".join(lines)))

    # Remove the dummy column created by Fidelity's malformed header.
    df = df.loc[:, ~df.columns.str.startswith("Unnamed:")].copy()

    lookup = build_lookup(df.columns)

    last_price_col = get_column(lookup, "last_price")
    print(f"Last price column: {last_price_col}")
    bid_col = get_column(lookup, "bid")
    ask_col = get_column(lookup, "ask")
    qty_col = get_column(lookup, "quantity")
    market_value_col = get_column(lookup, "market_value")
    # symbol_col = get_column(lookup, "symbol")

    if bid_col is None:
        print("Error: Could not locate Bid column.")
        return 1

    if ask_col is None:
        print("Error: Could not locate Ask column.")
        return 1

    last = pd.to_numeric(
        df[last_price_col]
        .astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False),
        errors="coerce",
    )
    bid = pd.to_numeric(df[bid_col].replace("--", pd.NA), errors="coerce")
    ask = pd.to_numeric(df[ask_col].replace("--", pd.NA), errors="coerce")

    if qty_col:
        quantity = pd.to_numeric(df[qty_col], errors="coerce").fillna(1)
    else:
        quantity = 1

    #
    # Calculations
    #

    mid = (bid + ask) / 2

    # If bid/ask are unavailable, fall back to the last traded price.
    mid = mid.fillna(last)

    df["Mid"] = mid
    df["Mid Value"] = mid * quantity * 100

    cash_mask = (
        df["Symbol"].fillna("").str.contains(
            "SPAXX|Pending Activity",
            case=False,
            regex=True,
        )
    )

    df.loc[cash_mask, "Mid Value"] = pd.to_numeric(
        df.loc[cash_mask, market_value_col]
        .astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False),
        errors="coerce",
    )

    midpoint_total = df["Mid Value"].sum(skipna=True)

    fidelity_total = None
    difference = None
    percent_difference = None

    if market_value_col:
        fidelity_total = (
            pd.to_numeric(
                df[market_value_col]
                .astype(str)
                .str.replace("$", "", regex=False)
                .str.replace(",", "", regex=False),
                errors="coerce",
            )
            .fillna(0)
            .sum()
        )

        difference = midpoint_total - fidelity_total

        if fidelity_total != 0:
            percent_difference = difference / fidelity_total * 100

    spread_pct = ((ask - bid) / ((ask + bid) / 2) * 100).mean(skipna=True)

    #
    # Write workbook
    #

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Open Positions")

    wb = load_workbook(output_path)
    ws = wb["Open Positions"]

    headers = {
        cell.value: index + 1
        for index, cell in enumerate(ws[1])
    }

    for column in (
        bid_col,
        ask_col,
        market_value_col,
        "Mid",
        "Mid Value",
    ):
        format_currency(ws, headers, column)

    wb.save(output_path)

    #
    # Summary
    #

    print()
    print("Portfolio Summary")
    print("-----------------")
    print(f"Positions:                {len(df):>15}")

    if fidelity_total is not None:
        print(f"Fidelity Market Value:    ${fidelity_total:>15,.2f}")

    print(f"Midpoint Market Value:    ${midpoint_total:>15,.2f}")

    if difference is not None:
        sign = "+" if difference >= 0 else ""
        print(
            f"Difference:               ${difference:>15,.2f} ({sign}{percent_difference:.2f}%)"
        )

    if pd.notna(spread_pct):
        print(f"Average Bid/Ask Spread:   {spread_pct:>15.2f}%")

    print(f"Output:                   {output_path}")
    print()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
